#!/usr/bin/env python3
"""
Brand Contact Enrichment Script
Reads brand list from xlsx, finds influencer/social marketing contacts via RocketReach,
and writes results back to the spreadsheet.

Usage:
  python enrich_contacts.py --api-key YOUR_ROCKETREACH_KEY --input brand_prospecting_list.xlsx

Or set ROCKETREACH_API_KEY as an environment variable.

Resume: If interrupted, re-run with the same output file. It skips brands that already have
a contact name populated.
"""

import argparse
import json
import os
import sys
import time
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

sys.stdout.reconfigure(line_buffering=True)

# ── RocketReach helpers ──────────────────────────────────────────────────

def rr_request(url, api_key, payload=None, method=None):
    """Make a RocketReach API request. Returns parsed JSON."""
    headers = {"Api-Key": api_key, "Content-Type": "application/json"}
    if payload:
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers=headers, method=method or "POST")
    else:
        req = urllib.request.Request(url, headers=headers, method=method or "GET")

    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))


def search_brand_contacts(brand_name, api_key):
    """Search RocketReach for marketing/influencer contacts at a brand.
    Returns list of profiles (no emails - search is free, no credits used).
    """
    search_url = "https://api.rocketreach.co/api/v2/person/search"

    # Try progressively broader title searches
    title_searches = [
        ["influencer marketing", "creator partnerships", "influencer partnerships", "creator marketing"],
        ["social media", "head of social", "social media manager", "social media director"],
        ["brand partnerships", "brand marketing", "partnerships manager", "marketing manager"],
    ]

    for titles in title_searches:
        try:
            data = rr_request(search_url, api_key, payload={
                "query": {
                    "current_employer": [f'"{brand_name}"'],
                    "current_title": titles,
                },
                "page_size": 5,
                "order_by": "popularity",
            })
            profiles = data.get("profiles", [])
            if profiles:
                return profiles
        except urllib.error.HTTPError as e:
            if e.code == 429:
                print(f"    Rate limited, waiting 10s...")
                time.sleep(10)
                continue
            body = ""
            try:
                body = e.read().decode()[:200]
            except Exception:
                pass
            print(f"    Search HTTP {e.code}: {body}")
            return []
        except Exception as e:
            print(f"    Search error: {e}")
            return []

    return []


def pick_best_profile(profiles):
    """From a list of RocketReach profiles, pick the best influencer/social contact.
    Prefer: influencer/creator in title > social in title > director/head/vp level > first result.
    """
    if not profiles:
        return None

    # Score each profile
    scored = []
    for p in profiles:
        title = (p.get("current_title") or "").lower()
        score = 0
        # Strong signals
        if "influencer" in title:
            score += 10
        if "creator" in title:
            score += 10
        if "partnership" in title:
            score += 5
        if "social" in title:
            score += 5
        # Seniority bonus
        if any(kw in title for kw in ["director", "head", "vp", "vice president"]):
            score += 3
        if any(kw in title for kw in ["senior", "manager", "lead"]):
            score += 2
        # Slight penalty for very senior (CMO/CEO probably not the right outreach target)
        if any(kw in title for kw in ["ceo", "cfo", "coo", "chief"]):
            score -= 2
        scored.append((score, p))

    scored.sort(key=lambda x: x[0], reverse=True)
    return scored[0][1]


def lookup_email(profile_id, api_key):
    """Lookup a person's email by RocketReach profile ID. Costs 1 export credit."""
    url = f"https://api.rocketreach.co/api/v2/person/lookup?id={profile_id}"
    try:
        data = rr_request(url, api_key, method="GET")
        status = data.get("status", "")

        if status == "complete":
            # Try to get professional email first
            emails = data.get("emails", [])
            if emails:
                work = [e for e in emails if e.get("type") == "professional"]
                if work:
                    return work[0].get("email")
                return emails[0].get("email")
            # Fallback fields
            return (
                data.get("current_work_email")
                or data.get("recommended_professional_email")
                or data.get("current_personal_email")
            )
        elif status == "searching":
            # Wait and retry once
            print(f"    Email lookup processing, waiting 5s...")
            time.sleep(5)
            data2 = rr_request(url, api_key, method="GET")
            if data2.get("status") == "complete":
                emails2 = data2.get("emails", [])
                if emails2:
                    work = [e for e in emails2 if e.get("type") == "professional"]
                    return work[0].get("email") if work else emails2[0].get("email")
            return None
        else:
            return None

    except urllib.error.HTTPError as e:
        if e.code == 429:
            print(f"    Rate limited on lookup, waiting 10s...")
            time.sleep(10)
            return None
        return None
    except Exception as e:
        print(f"    Lookup error: {e}")
        return None


# ── Main enrichment loop ─────────────────────────────────────────────────

def enrich_spreadsheet(input_path, output_path, api_key, max_brands=None, start_row=None):
    """Read brands from xlsx, enrich with contacts, save results."""

    wb = openpyxl.load_workbook(input_path)
    ws = wb['Brand Prospecting List']

    # Check if contact columns already exist
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    existing_cols = len(header_row)

    # Add new columns if not present
    new_headers = ["Contact Name", "Contact Title", "Contact Email", "LinkedIn URL", "Contact Source"]
    header_fill = PatternFill('solid', fgColor='1A1A2E')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)

    if "Contact Name" not in header_row:
        for i, h in enumerate(new_headers):
            col = existing_cols + 1 + i
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='left', vertical='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(existing_cols + 1)].width = 22
        ws.column_dimensions[openpyxl.utils.get_column_letter(existing_cols + 2)].width = 30
        ws.column_dimensions[openpyxl.utils.get_column_letter(existing_cols + 3)].width = 32
        ws.column_dimensions[openpyxl.utils.get_column_letter(existing_cols + 4)].width = 40
        ws.column_dimensions[openpyxl.utils.get_column_letter(existing_cols + 5)].width = 16
        contact_col_start = existing_cols + 1
    else:
        contact_col_start = header_row.index("Contact Name") + 1

    # Counters
    total = ws.max_row - 1
    processed = 0
    found = 0
    emails_found = 0
    skipped = 0
    exports_used = 0
    begin_row = start_row or 2

    print(f"\n{'='*60}")
    print(f"Brand Contact Enrichment")
    print(f"{'='*60}")
    print(f"Total brands: {total}")
    print(f"Starting at row: {begin_row}")
    if max_brands:
        print(f"Max brands this run: {max_brands}")
    print(f"{'='*60}\n")

    for row in range(begin_row, ws.max_row + 1):
        brand_name = ws.cell(row=row, column=1).value
        category = ws.cell(row=row, column=2).value
        existing_contact = ws.cell(row=row, column=contact_col_start).value

        if not brand_name:
            continue

        # Skip if already enriched
        if existing_contact and str(existing_contact).strip():
            skipped += 1
            continue

        if max_brands and processed >= max_brands:
            print(f"\nReached max brands limit ({max_brands}). Saving progress...")
            break

        processed += 1
        print(f"[{processed}/{total - skipped}] {brand_name} ({category})")

        # Step 1: Search (free)
        profiles = search_brand_contacts(brand_name, api_key)

        if not profiles:
            print(f"    No marketing contacts found")
            ws.cell(row=row, column=contact_col_start, value="NOT FOUND")
            ws.cell(row=row, column=contact_col_start + 4, value="rocketreach")
            # Save every 25 brands
            if processed % 25 == 0:
                wb.save(output_path)
                print(f"    [saved progress]")
            time.sleep(0.5)  # Be nice to the API
            continue

        # Step 2: Pick best match
        best = pick_best_profile(profiles)
        if not best:
            ws.cell(row=row, column=contact_col_start, value="NOT FOUND")
            continue

        name = best.get("name", "")
        title = best.get("current_title", "")
        linkedin = best.get("linkedin_url", "")
        profile_id = best.get("id")

        found += 1
        print(f"    Found: {name} - {title}")

        # Step 3: Lookup email (costs 1 export)
        email = None
        if profile_id:
            email = lookup_email(profile_id, api_key)
            exports_used += 1
            if email:
                emails_found += 1
                print(f"    Email: {email}")
            else:
                print(f"    No email returned")

        # Write to spreadsheet
        data_font = Font(name='Arial', size=10)
        ws.cell(row=row, column=contact_col_start, value=name).font = data_font
        ws.cell(row=row, column=contact_col_start + 1, value=title).font = data_font
        ws.cell(row=row, column=contact_col_start + 2, value=email or "").font = data_font
        ws.cell(row=row, column=contact_col_start + 3, value=linkedin or "").font = data_font
        ws.cell(row=row, column=contact_col_start + 4, value="rocketreach").font = data_font

        # Save every 25 brands
        if processed % 25 == 0:
            wb.save(output_path)
            print(f"    [saved progress - {found} contacts, {emails_found} emails, {exports_used} exports used]")

        # Rate limiting - 0.5s between brands
        time.sleep(0.5)

    # Final save
    wb.save(output_path)

    print(f"\n{'='*60}")
    print(f"RESULTS")
    print(f"{'='*60}")
    print(f"Processed:     {processed}")
    print(f"Skipped:       {skipped} (already enriched)")
    print(f"Contacts found: {found}")
    print(f"Emails found:  {emails_found}")
    print(f"Exports used:  {exports_used}")
    print(f"Output:        {output_path}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Enrich brand list with contacts via RocketReach")
    parser.add_argument("--api-key", default=os.environ.get("ROCKETREACH_API_KEY"), help="RocketReach API key")
    parser.add_argument("--input", default="brand_prospecting_list.xlsx", help="Input xlsx file")
    parser.add_argument("--output", default=None, help="Output xlsx file (defaults to input file)")
    parser.add_argument("--max", type=int, default=None, help="Max brands to process this run")
    parser.add_argument("--start-row", type=int, default=None, help="Start at this row (2 = first brand)")
    args = parser.parse_args()

    if not args.api_key:
        print("ERROR: No API key. Set ROCKETREACH_API_KEY or use --api-key")
        sys.exit(1)

    output = args.output or args.input
    enrich_contacts = enrich_spreadsheet(args.input, output, args.api_key, args.max, args.start_row)
